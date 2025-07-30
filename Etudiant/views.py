# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Avg, Count
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.models import User
import pandas as pd
import json
from io import BytesIO


from .models import Classe, Etudiant, Matiere, Note, Bulletin
from .forms import (
    ClasseForm, EtudiantForm, MatiereForm, NoteForm, NoteRapideForm,
    ImportDonneesForm, RechercheEtudiantForm, GenerationBulletinForm
)

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from io import BytesIO
import zipfile
from datetime import datetime
from decimal import Decimal
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from utilisateurs.models import ProfilUtilisateur
from django.http import HttpResponseForbidden
# ================= VUES GÉNÉRALES =================


@login_required
def dashboard(request):
    """Vue du tableau de bord principal filtré par compte"""
    # Récupère le compte lié à l'utilisateur connecté
    e = Etudiant.objects.all().count()
    print(e)
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé. Contacte l'administrateur.")
    compte = profil.compte

    # Statistiques générales filtrées par compte
    total_etudiants = Etudiant.objects.filter(actif=True, compte=compte).count()
    total_classes = Classe.objects.filter(compte=compte).count()
    total_matieres = Matiere.objects.filter(actif=True, compte=compte).count()
    total_notes = Note.objects.filter(compte=compte).count()
    
    # Étudiants récents du compte
    etudiants_recents = Etudiant.objects.filter(actif=True, compte=compte).order_by('-date_inscription')[:5]
    
    # Notes récentes du compte
    notes_recentes = Note.objects.select_related('etudiant', 'matiere').filter(compte=compte).order_by('-date_saisie')[:10]
    
    context = {
        'total_etudiants': total_etudiants,
        'total_classes': total_classes,
        'total_matieres': total_matieres,
        'total_notes': total_notes,
        'etudiants_recents': etudiants_recents,
        'notes_recentes': notes_recentes,
    }
    return render(request, 'gestion/dashboard.html', context)

# ================= GESTION DES CLASSES =================

@login_required
def liste_classes(request):
    """Liste des classes, filtrées par compte"""
    # Récupère le compte de l'utilisateur connecté
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé. Contacte l'administrateur.")
    compte = profil.compte

    # Filtrage des classes liées au compte
    classes = Classe.objects.filter(compte=compte).annotate(
        nb_etudiants=Count('etudiant', filter=Q(etudiant__actif=True))
    ).order_by('-annee_scolaire', 'niveau', 'nom')
    
    # Pagination
    paginator = Paginator(classes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'gestion/classes/liste.html', {
        'page_obj': page_obj
    })


@login_required
def ajouter_classe(request):
    """Ajouter une nouvelle classe"""

    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé. Contacte l'administrateur.")

    compte = profil.compte  # récupération correcte du compte

    if request.method == 'POST':
        form = ClasseForm(request.POST)
        if form.is_valid():
            classe = form.save(commit=False)
            classe.compte = compte  # on lie à l’utilisateur via son compte
            classe.save()
            messages.success(request, 'Classe ajoutée avec succès!')
            return redirect('liste_classes')
    else:
        form = ClasseForm()

    return render(request, 'gestion/classes/ajouter.html', {'form': form})

@login_required
def modifier_classe(request, pk):
    """Modifier une classe"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé. Contacte l'administrateur.")

    compte = profil.compte
    classe = get_object_or_404(Classe, pk=pk)

    # Vérifie que la classe appartient bien à ce compte
    if classe.compte != compte:
        return HttpResponseForbidden("Vous n'avez pas le droit de modifier cette classe.")

    if request.method == 'POST':
        form = ClasseForm(request.POST, instance=classe)
        if form.is_valid():
            form.save()
            messages.success(request, 'Classe modifiée avec succès!')
            return redirect('liste_classes')
    else:
        form = ClasseForm(instance=classe)

    return render(request, 'gestion/classes/modifier.html', {
        'form': form, 'classe': classe
    })

@login_required
def supprimer_classe(request, pk):
    """Supprimer une classe"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé. Contacte l'administrateur.")

    compte = profil.compte
    classe = get_object_or_404(Classe, pk=pk)

    # Vérifie que la classe appartient bien à ce compte
    if classe.compte != compte:
        return HttpResponseForbidden("Vous n'avez pas le droit de supprimer cette classe.")

    if request.method == 'POST':
        classe.delete()
        messages.success(request, 'Classe supprimée avec succès!')
        return redirect('liste_classes')

    return render(request, 'gestion/classes/supprimer.html', {'classe': classe})

# ================= GESTION DES ÉTUDIANTS =================

@login_required
def liste_etudiants(request):
    """Liste des étudiants avec recherche, filtres et filtrage par compte utilisateur connecté"""
    
    # Récupérer le profil utilisateur pour accéder au compte
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé. Contacte l'administrateur.")
    
    compte = profil.compte

    form = RechercheEtudiantForm(request.GET or None, user=request.user)
    
    # On filtre uniquement les étudiants liés au compte de l'utilisateur connecté
    etudiants = Etudiant.objects.select_related('classe').filter(compte=compte)
    
    if form.is_valid():
        recherche = form.cleaned_data.get('recherche')
        classe = form.cleaned_data.get('classe')
        actif = form.cleaned_data.get('actif')
        
        if recherche:
            etudiants = etudiants.filter(
                Q(nom__icontains=recherche) |
                Q(prenom__icontains=recherche) |
                Q(numero_etudiant__icontains=recherche)
            )
        
        if classe:
            etudiants = etudiants.filter(classe=classe)
        
        if actif:
            etudiants = etudiants.filter(actif=actif == 'True')
    
    etudiants = etudiants.order_by('nom', 'prenom')
    
    paginator = Paginator(etudiants, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'gestion/etudiants/liste.html', {
        'page_obj': page_obj,
        'form': form
    })



@login_required
def ajouter_etudiant(request):
    """Ajouter un nouvel étudiant lié au compte de l'utilisateur connecté"""
    
    # Récupérer le profil utilisateur pour accéder au compte
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé. Contacte l'administrateur.")
    
    compte = profil.compte
    
    if request.method == 'POST':
        form = EtudiantForm(request.POST or None, user=request.user )
        if form.is_valid():
            etudiant = form.save(commit=False)  # On crée l'instance sans enregistrer
            etudiant.compte = compte             # On assigne le compte lié à l'utilisateur
            etudiant.save()                      # On sauvegarde en base
            messages.success(request, 'Étudiant ajouté avec succès!')
            return redirect('liste_etudiants')
    else:
        form = EtudiantForm(user=request.user)
    
    return render(request, 'gestion/etudiants/ajouter.html', {'form': form})


@login_required
def modifier_etudiant(request, pk):
    """Modifier un étudiant lié au compte utilisateur connecté"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    compte = profil.compte

    # On s'assure que l'étudiant appartient au compte
    etudiant = get_object_or_404(Etudiant, pk=pk, compte=compte)

    if request.method == 'POST':
        form = EtudiantForm(request.POST, instance=etudiant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Étudiant modifié avec succès!')
            return redirect('liste_etudiants')
    else:
        form = EtudiantForm(instance=etudiant)
    
    return render(request, 'gestion/etudiants/modifier.html', {
        'form': form,
        'etudiant': etudiant
    })

@login_required
def detail_etudiant(request, pk):
    """Détail d'un étudiant avec ses notes, accessible uniquement si lié au compte"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    compte = profil.compte

    etudiant = get_object_or_404(Etudiant, pk=pk, compte=compte)

    notes = Note.objects.filter(etudiant=etudiant).select_related('matiere').order_by('-date_evaluation')
    moyenne_generale = notes.aggregate(avg=Avg('note'))['avg']

    return render(request, 'gestion/etudiants/detail.html', {
        'etudiant': etudiant,
        'notes': notes,
        'moyenne_generale': moyenne_generale
    })

@login_required
def supprimer_etudiant(request, pk):
    """Supprimer un étudiant, uniquement si lié au compte utilisateur connecté"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    compte = profil.compte

    etudiant = get_object_or_404(Etudiant, pk=pk, compte=compte)

    if request.method == 'POST':
        etudiant.delete()
        messages.success(request, 'Étudiant supprimé avec succès!')
        return redirect('liste_etudiants')

    return render(request, 'gestion/etudiants/supprimer.html', {'etudiant': etudiant})
# ================= GESTION DES MATIÈRES =================

@login_required
def liste_matieres(request):
    """Liste des matières liées au compte de l'utilisateur connecté"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    compte = profil.compte

    matieres = Matiere.objects.select_related('enseignant').annotate(
        nb_notes=Count('note')
    ).filter(compte=compte).order_by('nom')

    return render(request, 'gestion/matieres/liste.html', {'matieres': matieres})

@login_required
def ajouter_matiere(request):
    """Ajouter une nouvelle matière liée au compte utilisateur connecté"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    compte = profil.compte

    if request.method == 'POST':
        form = MatiereForm(request.POST, user=request.user)
        if form.is_valid():
            matiere = form.save(commit=False)
            # Associer l'enseignant et/ou le compte à la matière selon ton modèle
            matiere.compte = compte  # ou matiere.compte = compte, à adapter
            matiere.save()
            messages.success(request, 'Matière ajoutée avec succès!')
            return redirect('liste_matieres')
    else:
        form = MatiereForm(user=request.user)
    
    return render(request, 'gestion/matieres/ajouter.html', {'form': form})

@login_required
def modifier_matiere(request, pk):
    """Modifier une matière, uniquement si liée au compte utilisateur connecté"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    compte = profil.compte

    # On récupère la matière liée à ce compte (par exemple via enseignant__compte)
    matiere = get_object_or_404(Matiere, pk=pk, compte=compte)

    if request.method == 'POST':
        form = MatiereForm(request.POST, instance=matiere, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Matière modifiée avec succès!')
            return redirect('liste_matieres')
    else:
        form = MatiereForm(instance=matiere, user=request.user)
    
    return render(request, 'gestion/matieres/modifier.html', {
        'form': form,
        'matiere': matiere
    })

@login_required
def supprimer_matiere(request, pk):
    """Supprimer une matière, uniquement si liée au compte utilisateur connecté"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    compte = profil.compte

    matiere = get_object_or_404(Matiere, pk=pk, compte=compte)

    if request.method == 'POST':
        matiere.delete()
        messages.success(request, 'Matière supprimée avec succès!')
        return redirect('liste_matieres')
    
    return render(request, 'gestion/matieres/supprimer.html', {'matiere': matiere})


# ================= GESTION DES NOTES =================



@login_required
def liste_notes(request):
    """Liste des notes liées au compte utilisateur connecté"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    compte = profil.compte

    # Filtrer uniquement les notes liées au compte (via les étudiants du compte)
    notes = Note.objects.select_related('etudiant', 'matiere', 'modifie_par').filter(
        compte=compte  # adapte ici si la relation est différente
    ).order_by('-date_evaluation')

    # Filtres supplémentaires
    matiere_id = request.GET.get('matiere')
    classe_id = request.GET.get('classe')
    
    if matiere_id:
        notes = notes.filter(matiere_id=matiere_id)
    
    if classe_id:
        notes = notes.filter(etudiant__classe_id=classe_id)

    paginator = Paginator(notes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Pour les filtres : uniquement les matières et classes du compte
    matieres = Matiere.objects.filter(compte=compte, actif=True).distinct()
    classes = Classe.objects.filter(compte=compte)

    return render(request, 'gestion/notes/liste.html', {
        'page_obj': page_obj,
        'matieres': matieres,
        'classes': classes
    })
    
@login_required
def ajouter_note(request):
    """Ajouter une note liée au compte utilisateur connecté"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    compte = profil.compte

    if request.method == 'POST':
        form = NoteForm(request.POST,user=request.user)
        if form.is_valid():
            note = form.save(commit=False)
            note.modifie_par = request.user
            # Sécurité : on vérifie que la note est liée à un étudiant du bon compte
            note.compte = compte  # ou matiere.compte = compte, à adapter
          
            note.save()
            messages.success(request, 'Note ajoutée avec succès!')
            return redirect('liste_notes')
    else:
        form = NoteForm(user=request.user)
    
    return render(request, 'gestion/notes/ajouter.html', {'form': form})

@login_required
def modifier_note(request, pk):
    """Modifier une note uniquement si liée au compte utilisateur connecté"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    compte = profil.compte

    note = get_object_or_404(Note, pk=pk, compte=compte)

    if request.method == 'POST':
        form = NoteForm(request.POST, instance=note, user=request.user)
        if form.is_valid():
            note = form.save(commit=False)
            note.modifie_par = request.user
            note.save()
            messages.success(request, 'Note modifiée avec succès!')
            return redirect('liste_notes')
    else:
        form = NoteForm(instance=note, user=request.user)
    
    return render(request, 'gestion/notes/modifier.html', {
        'form': form,
        'matiere': note
    })


@login_required
def supprimer_note(request, pk):
    """Supprimer une note uniquement si liée au compte utilisateur connecté"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    compte = profil.compte

    note = get_object_or_404(Note, pk=pk, compte=compte)

    if request.method == 'POST':
        note.delete()
        messages.success(request, 'Note supprimée avec succès!')
        return redirect('liste_notes')
    
    return render(request, 'gestion/notes/supprimer.html', {'note': note})


@login_required
def saisie_rapide_notes(request):
    """Saisie rapide de notes pour une classe liée au compte utilisateur connecté"""
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé.")
    
    compte = profil.compte
    classe_id = request.GET.get('classe')

    if request.method == 'POST':
        form = NoteRapideForm(request.POST, user=request.user)
        if form.is_valid() and classe_id:
            # Sécuriser l’accès à la classe
            classe = get_object_or_404(Classe, pk=classe_id, compte=compte)
            etudiants = classe.etudiant_set.filter(actif=True, compte=compte)

            matiere = form.cleaned_data['matiere']
            type_evaluation = form.cleaned_data['type_evaluation']
            date_evaluation = form.cleaned_data['date_evaluation']
            semestre = form.cleaned_data['semestre']
            note_sur = form.cleaned_data['note_sur']

            notes_ajoutees = 0
            for etudiant in etudiants:
                note_value = request.POST.get(f'note_{etudiant.id}')
                if note_value:
                    try:
                        note_value = float(note_value)
                        Note.objects.create(
                            etudiant=etudiant,
                            matiere=matiere,
                            note=note_value,
                            note_sur=note_sur,
                            type_evaluation=type_evaluation,
                            date_evaluation=date_evaluation,
                            semestre=semestre,
                            modifie_par=request.user
                        )
                        notes_ajoutees += 1
                    except ValueError:
                        pass

            messages.success(request, f'{notes_ajoutees} notes ajoutées avec succès!')
            return redirect('liste_notes')
    else:
        form = NoteRapideForm(user=request.user)

    # Récupération sécurisée des étudiants
    etudiants = []
    if classe_id:
        try:
            classe = Classe.objects.get(pk=classe_id, compte=compte)
            etudiants = classe.etudiant_set.filter(actif=True, compte=compte).order_by('nom', 'prenom')
        except Classe.DoesNotExist:
            return HttpResponseForbidden("Cette classe ne vous appartient pas.")

    classes = Classe.objects.filter(compte=compte)

    return render(request, 'gestion/notes/saisie_rapide.html', {
        'form': form,
        'classes': classes,
        'etudiants': etudiants,
        'classe_selectionnee': classe_id
    })


# ================= IMPORT/EXPORT =================

@login_required
def importer_donnees(request):
    # Récupérer le compte lié à l'utilisateur connecté
    try:
        profil = ProfilUtilisateur.objects.get(user=request.user)
        compte = profil.compte
    except ProfilUtilisateur.DoesNotExist:
        return HttpResponseForbidden("Aucun profil utilisateur associé. Contacte l'administrateur.")

    if request.method == 'POST':
        form = ImportDonneesForm(request.POST, request.FILES)
        if form.is_valid():
            type_import = form.cleaned_data['type_import']
            fichier = form.cleaned_data['fichier']

            print(f"Type d'import: {type_import}")
            print(f"Fichier reçu: {fichier.name}")

            try:
                # Lecture du fichier
                if fichier.name.endswith('.csv'):
                    df = pd.read_csv(fichier)
                else:
                    df = pd.read_excel(fichier)

                print(f"Colonnes du fichier: {df.columns.tolist()}")
                print(f"Nombre de lignes dans le fichier: {len(df)}")

                nb_importes = 0

                # ----------- Import Étudiants -----------
                if type_import == 'etudiants':
                    colonnes = [
                        'numero_etudiant', 'nom', 'prenom', 'date_naissance', 'sexe',
                        'adresse', 'telephone', 'email', 'classe_id'
                    ]
                    # Vérifier les colonnes nécessaires (sans compte_id)
                    for col in colonnes:
                        if col not in df.columns:
                            print(f"Colonne manquante : {col}")
                            messages.error(request, f"Colonne manquante : {col}")
                            return redirect('importer_donnees')

                    for index, row in df.iterrows():
                        print(f"Ligne {index} étudiant: {row.to_dict()}")
                        try:
                            numero = str(row['numero_etudiant']).strip()
                            if Etudiant.objects.filter(numero_etudiant=numero).exists():
                                print("Déjà existant, on saute.")
                                continue

                            classe_id = int(row['classe_id'])
                            if not Classe.objects.filter(id=classe_id).exists():
                                print(f"Classe ID {classe_id} n'existe pas, on saute cette ligne.")
                                continue

                            etu = Etudiant.objects.create(
                                numero_etudiant=numero,
                                nom=row['nom'],
                                prenom=row['prenom'],
                                date_naissance=row['date_naissance'],
                                sexe=row['sexe'],
                                adresse=row.get('adresse', ''),
                                telephone=row.get('telephone', ''),
                                email=row.get('email', ''),
                                classe_id=classe_id,
                                compte=compte,  # Utilise le compte de l'utilisateur connecté
                            )
                            nb_importes += 1
                            print(f"Étudiant importé avec ID {etu.id}.")
                        except Exception as e:
                            print(f"Erreur import étudiant à la ligne {index}: {e}")
                            messages.warning(request, f"Étudiant non importé : {e}")

                    messages.success(request, f"{nb_importes} étudiant(s) importé(s) avec succès.")

                # ----------- Import Matières -----------
                elif type_import == 'matieres':
                    colonnes = [
                        'nom', 'code', 'coefficient', 'description',
                        'enseignant_id', 'actif'
                    ]
                    # Vérifier les colonnes nécessaires (sans compte_id)
                    for col in colonnes:
                        if col not in df.columns:
                            print(f"Colonne manquante : {col}")
                            messages.error(request, f"Colonne manquante : {col}")
                            return redirect('importer_donnees')

                    for index, row in df.iterrows():
                        print(f"Ligne {index} matière: {row.to_dict()}")
                        try:
                            if Matiere.objects.filter(code=row['code']).exists():
                                print("Déjà existante, on saute.")
                                continue

                            enseignant_id = int(row['enseignant_id']) if pd.notna(row['enseignant_id']) else None

                            matiere = Matiere.objects.create(
                                nom=row['nom'],
                                code=row['code'],
                                coefficient=float(row['coefficient']),
                                description=row.get('description', ''),
                                enseignant_id=enseignant_id,
                                compte=compte,  # Utilise le compte de l'utilisateur connecté
                                actif=str(row['actif']).lower() in ['true', '1']
                            )
                            nb_importes += 1
                            print(f"Matière importée avec ID {matiere.id}.")
                        except Exception as e:
                            print(f"Erreur import matière à la ligne {index}: {e}")
                            messages.warning(request, f"Matière non importée (code {row.get('code')}) : {e}")

                    messages.success(request, f"{nb_importes} matière(s) importée(s) avec succès.")

                # ----------- Import Classes -----------
                elif type_import == 'classe':
                    colonnes = ['nom', 'niveau', 'annee_scolaire']
                    # Vérifier les colonnes nécessaires (sans compte_id)
                    for col in colonnes:
                        if col not in df.columns:
                            print(f"Colonne manquante : {col}")
                            messages.error(request, f"Colonne manquante : {col}")
                            return redirect('importer_donnees')

                    for index, row in df.iterrows():
                        print(f"Ligne {index} classe: {row.to_dict()}")
                        try:
                            if Classe.objects.filter(
                                nom=row['nom'],
                                niveau=row['niveau'],
                                annee_scolaire=row['annee_scolaire'],
                                compte=compte
                            ).exists():
                                print("Classe déjà existante, on saute.")
                                continue

                            classe = Classe.objects.create(
                                nom=row['nom'],
                                niveau=row['niveau'],
                                annee_scolaire=row['annee_scolaire'],
                                compte=compte  # Utilise le compte de l'utilisateur connecté
                            )
                            nb_importes += 1
                            print(f"Classe importée avec ID {classe.id}.")
                        except Exception as e:
                            print(f"Erreur import classe à la ligne {index}: {e}")
                            messages.warning(request, f"Classe non importée (nom {row.get('nom')}) : {e}")

                    messages.success(request, f"{nb_importes} classe(s) importée(s) avec succès.")

                else:
                    print("Type d'importation non reconnu.")
                    messages.error(request, "Type d'importation non reconnu.")

                return redirect('importer_donnees')

            except Exception as e:
                print(f"Erreur générale: {e}")
                messages.error(request, f"Erreur lors du traitement du fichier : {e}")
                return redirect('importer_donnees')

    else:
        form = ImportDonneesForm()

    return render(request, 'gestion/import_export/import.html', {'form': form})







@login_required
def generation_bulletins(request):
    """Génération de bulletins de notes"""
    if request.method == 'POST':
        form = GenerationBulletinForm(request.POST)
        if form.is_valid():
            classe = form.cleaned_data['classe']
            semestre = form.cleaned_data['semestre']
            annee_scolaire = form.cleaned_data['annee_scolaire']
            format_export = form.cleaned_data['format_export']
            
            # Logique de génération des bulletins
            try:
                if format_export == 'pdf':
                    return generer_bulletins_pdf_individuels(classe, semestre, annee_scolaire, request.user)
                elif format_export == 'pdf_groupe':
                    return generer_bulletins_pdf_groupe(classe, semestre, annee_scolaire, request.user)
                elif format_export == 'excel':
                    return generer_bulletins_excel(classe, semestre, annee_scolaire, request.user)
                
                messages.success(request, 'Bulletins générés avec succès!')
            except Exception as e:
                messages.error(request, f'Erreur lors de la génération: {str(e)}')
    else:
        form = GenerationBulletinForm()
    
    return render(request, 'gestion/import_export/bulletins.html', {'form': form})

def generer_bulletins_pdf_individuels(classe, semestre, annee_scolaire, user):
    """Génération de bulletins PDF individuels dans un ZIP"""
    
    # Récupérer tous les étudiants de la classe
    etudiants = Etudiant.objects.filter(classe=classe, actif=True)
    
    if not etudiants.exists():
        raise Exception("Aucun étudiant trouvé dans cette classe")
    
    # Créer un fichier ZIP en mémoire
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for etudiant in etudiants:
            # Générer le PDF pour chaque étudiant
            pdf_buffer = generer_bulletin_etudiant_pdf(etudiant, semestre, annee_scolaire, user)
            
            # Nom du fichier PDF
            filename = f"bulletin_{etudiant.nom}_{etudiant.prenom}_{semestre}_{annee_scolaire}.pdf"
            filename = filename.replace(' ', '_').replace('/', '-')
            
            # Ajouter le PDF au ZIP
            zip_file.writestr(filename, pdf_buffer.getvalue())
    
    # Préparer la réponse HTTP
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="bulletins_{classe.nom}_{semestre}_{annee_scolaire}.zip"'
    
    return response

def generer_bulletins_pdf_groupe(classe, semestre, annee_scolaire, user):
    """Génération d'un PDF groupé avec tous les bulletins"""
    
    etudiants = Etudiant.objects.filter(classe=classe, actif=True)
    
    if not etudiants.exists():
        raise Exception("Aucun étudiant trouvé dans cette classe")
    
    # Créer le PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    
    # Styles
    styles = getSampleStyleSheet()
    story = []
    
    for i, etudiant in enumerate(etudiants):
        if i > 0:  # Saut de page entre chaque bulletin
            story.append(Spacer(1, 20*cm))  # Force un saut de page
        
        # Générer le contenu du bulletin pour cet étudiant
        bulletin_content = generer_contenu_bulletin(etudiant, semestre, annee_scolaire, styles)
        story.extend(bulletin_content)
    
    # Construire le PDF
    doc.build(story)
    buffer.seek(0)
    
    # Préparer la réponse HTTP
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="bulletins_groupe_{classe.nom}_{semestre}_{annee_scolaire}.pdf"'
    
    return response

def generer_bulletin_etudiant_pdf(etudiant, semestre, annee_scolaire, user):
    """Génère le PDF d'un bulletin individuel"""
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    
    # Styles
    styles = getSampleStyleSheet()
    story = []
    
    # Générer le contenu du bulletin
    bulletin_content = generer_contenu_bulletin(etudiant, semestre, annee_scolaire, styles)
    story.extend(bulletin_content)
    
    # Construire le PDF
    doc.build(story)
    buffer.seek(0)
    
    return buffer

def generer_contenu_bulletin(etudiant, semestre, annee_scolaire, styles):
    """Génère le contenu d'un bulletin"""
    
    story = []
    
    # Style personnalisé pour le titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1,  # Centré
        textColor=colors.darkblue
    )
    
    # En-tête du bulletin
    story.append(Paragraph("BULLETIN DE NOTES", title_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Informations de l'étudiant
    info_data = [
        ['Nom complet:', etudiant.nom_complet],
        ['Numéro étudiant:', etudiant.numero_etudiant],
        ['Classe:', str(etudiant.classe)],
        ['Semestre:', semestre],
        ['Année scolaire:', annee_scolaire],
    ]
    
    info_table = Table(info_data, colWidths=[4*cm, 8*cm])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 1*cm))
    
    # Récupérer les notes de l'étudiant pour ce semestre
    notes = Note.objects.filter(
        etudiant=etudiant,
        semestre=semestre
    ).select_related('matiere')
    
    if notes.exists():
        # Tableau des notes
        notes_data = [['Matière', 'Code', 'Note', 'Note/20', 'Type', 'Coefficient']]
        
        total_points = Decimal('0')
        total_coefficients = Decimal('0')
        
        for note in notes:
            note_sur_20 = Decimal(str(note.note_sur_vingt))
            coeff = note.matiere.coefficient
            
            total_points += note_sur_20 * coeff
            total_coefficients += coeff
            
            notes_data.append([
                note.matiere.nom,
                note.matiere.code,
                f"{note.note}/{note.note_sur}",
                f"{float(note_sur_20):.2f}",
                note.get_type_evaluation_display(),
                str(note.matiere.coefficient)
            ])
        
        # Calculer la moyenne
        moyenne = total_points / total_coefficients if total_coefficients > 0 else Decimal('0')
        moyenne = float(moyenne)  # Convertir en float pour l'affichage
        
        # Ajouter la ligne de moyenne
        notes_data.append(['', '', '', '', 'MOYENNE GÉNÉRALE', f"{moyenne:.2f}/20"])
        
        notes_table = Table(notes_data, colWidths=[4*cm, 2*cm, 2*cm, 2*cm, 3*cm, 2*cm])
        notes_table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            
            # Corps du tableau
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Ligne de moyenne
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        story.append(Paragraph("DÉTAIL DES NOTES", styles['Heading2']))
        story.append(Spacer(1, 0.3*cm))
        story.append(notes_table)
        story.append(Spacer(1, 1*cm))
        
        # Créer ou mettre à jour le bulletin
        bulletin, created = Bulletin.objects.get_or_create(
            etudiant=etudiant,
            semestre=semestre,
            annee_scolaire=annee_scolaire,
            defaults={
                'moyenne_generale': moyenne,
                'genere_par': None  # Vous pouvez passer l'utilisateur ici
            }
        )
        
        if not created:
            bulletin.moyenne_generale = moyenne
            bulletin.save()
        
        # Appréciation
        if moyenne >= 16:
            appreciation = "Très bien - Félicitations"
        elif moyenne >= 14:
            appreciation = "Bien - Continue ainsi"
        elif moyenne >= 12:
            appreciation = "Assez bien - Peut mieux faire"
        elif moyenne >= 10:
            appreciation = "Passable - Doit faire des efforts"
        else:
            appreciation = "Insuffisant - Beaucoup d'efforts nécessaires"
        
        story.append(Paragraph("APPRÉCIATION GÉNÉRALE", styles['Heading2']))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(appreciation, styles['Normal']))
        
    else:
        story.append(Paragraph("Aucune note trouvée pour ce semestre.", styles['Normal']))
    
    # Pied de page
    story.append(Spacer(1, 2*cm))
    date_str = datetime.now().strftime("%d/%m/%Y")
    story.append(Paragraph(f"Bulletin généré le {date_str}", styles['Normal']))
    
    return story
# ================= FONCTIONS UTILITAIRES =================



def generer_bulletins_excel(classe, semestre, annee_scolaire, user):
    """Génération de bulletins Excel avec plusieurs onglets"""
    
    # Récupérer tous les étudiants de la classe
    etudiants = Etudiant.objects.filter(classe=classe, actif=True).order_by('nom', 'prenom')
    
    if not etudiants.exists():
        raise Exception("Aucun étudiant trouvé dans cette classe")
    
    # Créer le workbook Excel
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # Styles Excel
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1F4E79")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 1. Onglet récapitulatif
    ws_recap = wb.create_sheet("Récapitulatif")
    
    # Titre de l'onglet récapitulatif
    ws_recap['A1'] = f"RÉCAPITULATIF - {classe.nom}"
    ws_recap['A1'].font = title_font
    ws_recap['A2'] = f"Semestre: {semestre} - Année: {annee_scolaire}"
    ws_recap.merge_cells('A1:H1')
    ws_recap.merge_cells('A2:H2')
    
    # En-têtes du récapitulatif
    headers_recap = ['N°', 'Nom', 'Prénom', 'Numéro Étudiant', 'Moyenne Générale', 'Rang', 'Mention', 'Nb Notes']
    
    for col, header in enumerate(headers_recap, 1):
        cell = ws_recap.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Calculer les moyennes pour le classement
    moyennes_etudiants = []
    
    for etudiant in etudiants:
        notes = Note.objects.filter(
            etudiant=etudiant,
            semestre=semestre
        ).select_related('matiere')
        
        if notes.exists():
            total_points = Decimal('0')
            total_coefficients = Decimal('0')
            
            for note in notes:
                note_sur_20 = Decimal(str(note.note_sur_vingt))
                coeff = note.matiere.coefficient
                total_points += note_sur_20 * coeff
                total_coefficients += coeff
            
            moyenne = float(total_points / total_coefficients) if total_coefficients > 0 else 0
        else:
            moyenne = 0
        
        moyennes_etudiants.append({
            'etudiant': etudiant,
            'moyenne': moyenne,
            'nb_notes': notes.count()
        })
    
    # Trier par moyenne décroissante pour le rang
    moyennes_etudiants.sort(key=lambda x: x['moyenne'], reverse=True)
    
    # Remplir le récapitulatif
    for i, data in enumerate(moyennes_etudiants, 1):
        etudiant = data['etudiant']
        moyenne = data['moyenne']
        nb_notes = data['nb_notes']
        
        # Déterminer la mention
        if moyenne >= 16:
            mention = "Très Bien"
        elif moyenne >= 14:
            mention = "Bien"
        elif moyenne >= 12:
            mention = "Assez Bien"  
        elif moyenne >= 10:
            mention = "Passable"
        else:
            mention = "Insuffisant"
        
        row = i + 4
        ws_recap.cell(row=row, column=1, value=i).border = border
        ws_recap.cell(row=row, column=2, value=etudiant.nom).border = border
        ws_recap.cell(row=row, column=3, value=etudiant.prenom).border = border
        ws_recap.cell(row=row, column=4, value=etudiant.numero_etudiant).border = border
        ws_recap.cell(row=row, column=5, value=round(moyenne, 2)).border = border
        ws_recap.cell(row=row, column=6, value=i).border = border  # Rang
        ws_recap.cell(row=row, column=7, value=mention).border = border
        ws_recap.cell(row=row, column=8, value=nb_notes).border = border
        
        # Créer ou mettre à jour le bulletin en base
        bulletin, created = Bulletin.objects.get_or_create(
            etudiant=etudiant,
            semestre=semestre,
            annee_scolaire=annee_scolaire,
            defaults={
                'moyenne_generale': moyenne,
                'rang': i,
                'effectif_classe': len(moyennes_etudiants),
                'genere_par': user
            }
        )
        
        if not created:
            bulletin.moyenne_generale = moyenne
            bulletin.rang = i
            bulletin.effectif_classe = len(moyennes_etudiants)
            bulletin.save()
    
    # Ajuster les largeurs des colonnes du récapitulatif
    ws_recap.column_dimensions['A'].width = 5
    ws_recap.column_dimensions['B'].width = 15
    ws_recap.column_dimensions['C'].width = 15
    ws_recap.column_dimensions['D'].width = 15
    ws_recap.column_dimensions['E'].width = 12
    ws_recap.column_dimensions['F'].width = 8
    ws_recap.column_dimensions['G'].width = 12
    ws_recap.column_dimensions['H'].width = 10
    
    # 2. Onglet détaillé par matière
    ws_detail = wb.create_sheet("Notes par Matière")
    
    # Titre de l'onglet détaillé
    ws_detail['A1'] = f"NOTES DÉTAILLÉES - {classe.nom}"
    ws_detail['A1'].font = title_font
    ws_detail['A2'] = f"Semestre: {semestre} - Année: {annee_scolaire}"
    ws_detail.merge_cells('A1:I1')
    ws_detail.merge_cells('A2:I2')
    
    # En-têtes du détail
    headers_detail = ['Étudiant', 'N° Étudiant', 'Matière', 'Code', 'Note', 'Note/20', 'Type', 'Date', 'Coefficient']
    
    for col, header in enumerate(headers_detail, 1):
        cell = ws_detail.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Remplir les notes détaillées
    row_detail = 5
    for etudiant in etudiants:
        notes = Note.objects.filter(
            etudiant=etudiant,
            semestre=semestre
        ).select_related('matiere').order_by('matiere__nom', 'date_evaluation')
        
        for note in notes:
            ws_detail.cell(row=row_detail, column=1, value=etudiant.nom_complet).border = border
            ws_detail.cell(row=row_detail, column=2, value=etudiant.numero_etudiant).border = border
            ws_detail.cell(row=row_detail, column=3, value=note.matiere.nom).border = border
            ws_detail.cell(row=row_detail, column=4, value=note.matiere.code).border = border
            ws_detail.cell(row=row_detail, column=5, value=f"{note.note}/{note.note_sur}").border = border
            ws_detail.cell(row=row_detail, column=6, value=round(float(note.note_sur_vingt), 2)).border = border
            ws_detail.cell(row=row_detail, column=7, value=note.get_type_evaluation_display()).border = border
            ws_detail.cell(row=row_detail, column=8, value=note.date_evaluation.strftime('%d/%m/%Y')).border = border
            ws_detail.cell(row=row_detail, column=9, value=float(note.matiere.coefficient)).border = border
            row_detail += 1
    
    # Ajuster les largeurs des colonnes du détail
    ws_detail.column_dimensions['A'].width = 20
    ws_detail.column_dimensions['B'].width = 15
    ws_detail.column_dimensions['C'].width = 20
    ws_detail.column_dimensions['D'].width = 8
    ws_detail.column_dimensions['E'].width = 10
    ws_detail.column_dimensions['F'].width = 10
    ws_detail.column_dimensions['G'].width = 15
    ws_detail.column_dimensions['H'].width = 12
    ws_detail.column_dimensions['I'].width = 10
    
    # 3. Onglets individuels pour chaque étudiant
    for etudiant in etudiants:
        ws_etudiant = wb.create_sheet(f"{etudiant.nom[:10]}_{etudiant.prenom[:10]}")
        
        # Informations de l'étudiant
        ws_etudiant['A1'] = "BULLETIN INDIVIDUEL"
        ws_etudiant['A1'].font = title_font
        ws_etudiant.merge_cells('A1:D1')
        
        ws_etudiant['A3'] = "Nom complet:"
        ws_etudiant['B3'] = etudiant.nom_complet
        ws_etudiant['A4'] = "N° Étudiant:"
        ws_etudiant['B4'] = etudiant.numero_etudiant
        ws_etudiant['A5'] = "Classe:"
        ws_etudiant['B5'] = str(etudiant.classe)
        ws_etudiant['A6'] = "Semestre:"
        ws_etudiant['B6'] = semestre
        ws_etudiant['A7'] = "Année:"
        ws_etudiant['B7'] = annee_scolaire
        
        # Mettre en gras les labels
        for row in range(3, 8):
            ws_etudiant.cell(row=row, column=1).font = Font(bold=True)
        
        # Notes de l'étudiant
        notes = Note.objects.filter(
            etudiant=etudiant,
            semestre=semestre
        ).select_related('matiere')
        
        if notes.exists():
            # En-têtes des notes
            headers_notes = ['Matière', 'Code', 'Note', 'Note/20', 'Type', 'Date', 'Coefficient']
            
            for col, header in enumerate(headers_notes, 1):
                cell = ws_etudiant.cell(row=9, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Remplir les notes
            row_notes = 10
            total_points = Decimal('0')
            total_coefficients = Decimal('0')
            
            for note in notes:
                note_sur_20 = Decimal(str(note.note_sur_vingt))
                coeff = note.matiere.coefficient
                
                total_points += note_sur_20 * coeff
                total_coefficients += coeff
                
                ws_etudiant.cell(row=row_notes, column=1, value=note.matiere.nom).border = border
                ws_etudiant.cell(row=row_notes, column=2, value=note.matiere.code).border = border
                ws_etudiant.cell(row=row_notes, column=3, value=f"{note.note}/{note.note_sur}").border = border
                ws_etudiant.cell(row=row_notes, column=4, value=round(float(note_sur_20), 2)).border = border
                ws_etudiant.cell(row=row_notes, column=5, value=note.get_type_evaluation_display()).border = border
                ws_etudiant.cell(row=row_notes, column=6, value=note.date_evaluation.strftime('%d/%m/%Y')).border = border
                ws_etudiant.cell(row=row_notes, column=7, value=float(note.matiere.coefficient)).border = border
                row_notes += 1
            
            # Ligne de moyenne
            moyenne = float(total_points / total_coefficients) if total_coefficients > 0 else 0
            
            ws_etudiant.cell(row=row_notes, column=1, value="MOYENNE GÉNÉRALE").font = Font(bold=True)
            ws_etudiant.cell(row=row_notes, column=4, value=round(moyenne, 2)).font = Font(bold=True)
            
            # Ajuster les largeurs des colonnes
            ws_etudiant.column_dimensions['A'].width = 20
            ws_etudiant.column_dimensions['B'].width = 8
            ws_etudiant.column_dimensions['C'].width = 10
            ws_etudiant.column_dimensions['D'].width = 10
            ws_etudiant.column_dimensions['E'].width = 15
            ws_etudiant.column_dimensions['F'].width = 12
            ws_etudiant.column_dimensions['G'].width = 10
        
        else:
            ws_etudiant['A9'] = "Aucune note trouvée pour ce semestre"
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Préparer la réponse HTTP
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="bulletins_{classe.nom}_{semestre}_{annee_scolaire}.xlsx"'
    
    return response

# ================= VUES AJAX =================

@login_required
def get_etudiants_classe(request):
    """Récupérer les étudiants d'une classe (pour AJAX)"""
    classe_id = request.GET.get('classe_id')
    if classe_id:
        etudiants = Etudiant.objects.filter(
            classe_id=classe_id, actif=True
        ).values('id', 'nom', 'prenom', 'numero_etudiant')
        return JsonResponse({'etudiants': list(etudiants)})
    return JsonResponse({'etudiants': []})
